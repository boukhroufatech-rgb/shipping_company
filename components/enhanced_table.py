"""
Table améliorée avec toolbar et fonctionnalités avancées
"""
import csv
import re
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from PyQt6.QtPrintSupport import QPrinter, QPrintPreviewDialog, QPrintDialog
from PyQt6.QtGui import QTextDocument
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableView, QToolBar,
    QHeaderView, QPushButton, QLineEdit, QLabel, QMenu, QApplication,
    QFileDialog, QMessageBox, QFrame, QSizePolicy
)
from PyQt6.QtCore import Qt, pyqtSignal, QSortFilterProxyModel, QSettings, QTimer
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QAction, QIcon, QColor, QPageLayout

from .status_filter import StatusFilter
from .export_dialog import ExportPreviewDialog
from utils.icon_manager import IconManager
from core.themes import THEMES
from modules.settings.service import SettingsService


class MultiFilterProxyModel(QSortFilterProxyModel):
    """Proxy model permettant de filtrer sur plusieurs colonnes simultanément"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.column_filters = {}
        self.general_filter = ""

    def lessThan(self, left, right):
        """Natural sort for numeric columns"""
        left_data = left.data()
        right_data = right.data()
        
        # Try natural sort for first column (usually N°)
        if left.column() == 0 and left_data and right_data:
            try:
                left_val = int(left_data)
                right_val = int(right_data)
                return left_val < right_val
            except (ValueError, TypeError):
                # Fall back to string comparison
                return str(left_data) < str(right_data)
        
        # Default string comparison
        return str(left_data) < str(right_data) if left_data and right_data else False

    def set_column_filter(self, column, text):
        if not text:
            self.column_filters.pop(column, None)
        else:
            self.column_filters[column] = text.lower()
        self.invalidateFilter()

    def set_general_filter(self, text):
        self.general_filter = text.lower()
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row, source_parent):
        # 1. Filtre général
        if self.general_filter:
            found = False
            for col in range(self.sourceModel().columnCount()):
                index = self.sourceModel().index(source_row, col, source_parent)
                if self.general_filter in str(self.sourceModel().data(index)).lower():
                    found = True
                    break
            if not found:
                return False

        # 2. Filtres par colonne
        for col, text in self.column_filters.items():
            index = self.sourceModel().index(source_row, col, source_parent)
            if text not in str(self.sourceModel().data(index)).lower():
                return False
        
        return True


class EnhancedTableView(QWidget):
    """
    Table avec toolbar intégré et fonctionnalités avancées.
    """
    
    # Signaux
    addClicked = pyqtSignal()
    editClicked = pyqtSignal(int)
    deleteClicked = pyqtSignal(int)
    restoreClicked = pyqtSignal(int)
    refreshClicked = pyqtSignal()
    selectionChanged = pyqtSignal(list)
    
    def __init__(self, parent=None, table_id: str = None):
        super().__init__(parent)
        self.table_id = table_id
        self.model = QStandardItemModel()
        self.proxy_model = MultiFilterProxyModel()
        self.proxy_model.setSourceModel(self.model)
        
        self._permanent_hidden_columns = []
        self._column_stretch_factors = {}
        self._auto_fit_enabled = True
        self._row_numbering_enabled = False
        self._header_align_map = {}  # [CUSTOM] 2026-04-03 - Aligned columns
        self._running_balance_col = -1 # Index de la colonne du solde
        self._initial_balance = 0.0
        
        # Initialiser le footer model
        self.footer_model = QStandardItemModel(1, 0)
        
        # Initialiser les réglages et couleurs
        self.settings_service = SettingsService()
        self.active_theme = self.settings_service.get_setting("active_theme", "emerald")
        self.accent_color = THEMES.get(self.active_theme, THEMES["emerald"])["colors"]["accent"]
        
        # Initialiser l'interface
        self._setup_ui()
        
        # Connecter les signaux
        self.model.rowsInserted.connect(self._on_data_changed)
        self.model.rowsRemoved.connect(self._on_data_changed)
        self.model.dataChanged.connect(self._on_data_changed)
        self.proxy_model.layoutChanged.connect(self._update_footer)
        
    def _setup_ui(self):
        """Configure l'interface utilisateur"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # 1. Toolbar
        self.toolbar = self._create_toolbar()
        layout.addWidget(self.toolbar)
        
        # 2. Barre de recherche générale
        search_container = QWidget()
        search_container.setObjectName("searchContainer")
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(8, 4, 8, 4)
        
        search_label = QLabel()
        search_label.setPixmap(IconManager.get_icon("search", self.accent_color, 16).pixmap(16, 16))
        search_label.setStyleSheet("background: transparent; border: none;")
        
        self.search_input = QLineEdit()
        self.search_input.setObjectName("searchInput")
        self.search_input.setPlaceholderText("Recherche rapide... (Ctrl+F)")
        self.search_input.textChanged.connect(self._on_general_search)
        
        # [CUSTOM] 2026-04-03 - Ajouter autocomplete (completer)
        self._setup_search_completer()
        
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)
        
        # [CUSTOM] 2026-04-03 - Résultats counter
        self.search_count_label = QLabel("")
        self.search_count_label.setStyleSheet("color: #7d8590; font-size: 12px; padding: 0 8px;")
        search_layout.addWidget(self.search_count_label)
        
        # [CUSTOM] 2026-04-03 - Navigation buttons
        self.search_prev_btn = QPushButton("▲")
        self.search_prev_btn.setFixedSize(24, 24)
        self.search_prev_btn.setToolTip("Précédent (F3)")
        self.search_prev_btn.clicked.connect(self._search_prev)
        self.search_prev_btn.setStyleSheet("QPushButton { border: none; color: #7d8590; } QPushButton:hover { color: #2ea043; }")
        
        self.search_next_btn = QPushButton("▼")
        self.search_next_btn.setFixedSize(24, 24)
        self.search_next_btn.setToolTip("Suivant (F4)")
        self.search_next_btn.clicked.connect(self._search_next)
        self.search_next_btn.setStyleSheet("QPushButton { border: none; color: #7d8590; } QPushButton:hover { color: #2ea043; }")
        
        search_layout.addWidget(self.search_prev_btn)
        search_layout.addWidget(self.search_next_btn)
        
        search_layout.addStretch()
        layout.addWidget(search_container)
        
        # 3. Table principale
        self.table = QTableView()
        self.table.setObjectName("mainTable")
        self.table.setModel(self.proxy_model)
        self.table.setSortingEnabled(True)
        self.table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        self.table.sortByColumn(0, Qt.SortOrder.AscendingOrder)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.selectionModel().selectionChanged.connect(self._on_selection_changed)
        self.table.doubleClicked.connect(self._on_row_double_clicked)
        layout.addWidget(self.table)
        
        # 4. Table Footer
        self.footer = QTableView()
        self.footer.setObjectName("footerTable")
        self.footer.setFixedHeight(30)
        self.footer.setModel(self.footer_model)
        self.footer.horizontalHeader().setVisible(False)
        self.footer.verticalHeader().setVisible(False)
        self.footer.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.footer.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.footer.setFrameShape(QFrame.Shape.NoFrame)
        self.footer.setEditTriggers(QTableView.EditTrigger.NoEditTriggers)
        self.footer.setSelectionMode(QTableView.SelectionMode.NoSelection)
        layout.addWidget(self.footer)
        
        # 5. Synchronisation
        self.table.horizontalScrollBar().valueChanged.connect(self.footer.horizontalScrollBar().setValue)
        self.table.horizontalHeader().sectionResized.connect(self._sync_column_widths)
        
        # 6. Status Filter
        self.status_filter = StatusFilter()
        layout.addWidget(self.status_filter)

    def _on_general_search(self, text):
        self.proxy_model.set_general_filter(text)
        self._update_footer()
        self._update_completer(text)
        self._update_search_count()
        self._highlight_search_matches(text)
    
    def _highlight_search_matches(self, text: str):
        """[CUSTOM] 2026-04-08 - Highlight search text matches in yellow"""
        if not text:
            self._clear_search_highlight()
            return
        
        theme = THEMES.get(self.active_theme, THEMES["emerald"])
        text_primary = theme["colors"].get("text_primary", "#e6edf3")
        search_highlight_color = self.settings_service.get_setting("color_search_highlight", "#ffe000")
        search_text = text.lower()
        
        for row in range(self.model.rowCount()):
            for col in range(self.model.columnCount()):
                idx = self.model.index(row, col)
                if idx.isValid():
                    cell_text = str(idx.data() or "").lower()
                    
                    if search_text in cell_text:
                        item = self.model.item(row, col)
                        if item:
                            font = item.font()
                            font.setBold(True)
                            item.setFont(font)
                            item.setBackground(QColor(search_highlight_color))  # User-configured color
                            item.setForeground(QColor("#000000"))  # Black text for contrast
                    else:
                        item = self.model.item(row, col)
                        if item:
                            item.setBackground(QColor("transparent"))
                            item.setForeground(QColor(text_primary))
    
    def _clear_search_highlight(self):
        """Clear all search highlights"""
        theme = THEMES.get(self.active_theme, THEMES["emerald"])
        text_primary = theme["colors"].get("text_primary", "#e6edf3")
        
        for row in range(self.model.rowCount()):
            for col in range(self.model.columnCount()):
                item = self.model.item(row, col)
                if item:
                    item.setBackground(QColor("transparent"))
                    item.setForeground(QColor(text_primary))
                    font = item.font()
                    font.setBold(False)
                    item.setFont(font)
    
    def _update_search_count(self):
        """[CUSTOM] 2026-04-03 - Update results counter"""
        total = self.model.rowCount()
        filtered = self.proxy_model.rowCount()
        
        search_text = self.search_input.text()
        if search_text:
            self.search_count_label.setText(f"{filtered}/{total}")
        else:
            self.search_count_label.setText("")
    
    def _search_next(self):
        """[CUSTOM] 2026-04-03 - Navigate to next result"""
        self._search_navigate(1)
    
    def _search_prev(self):
        """[CUSTOM] 2026-04-03 - Navigate to previous result"""
        self._search_navigate(-1)
    
    def _search_navigate(self, direction: int):
        """Navigate through search results"""
        if not self.search_input.text():
            return
        
        rows = []
        search_text = self.search_input.text().lower()
        
        for row in range(self.proxy_model.rowCount()):
            for col in range(self.model.columnCount()):
                idx = self.proxy_model.index(row, col)
                if idx.isValid():
                    text = str(idx.data()).lower()
                    if search_text in text:
                        rows.append(row)
                        break
        
        if not rows:
            return
        
        # Get current row
        current = self.table.currentIndex()
        current_row = current.row() if current.isValid() else -1
        
        # Find next/prev
        if current_row not in rows:
            current_row = rows[0] - 1
        
        # Calculate next row
        current_idx = rows.index(current_row) if current_row in rows else -1
        next_idx = (current_idx + direction) % len(rows)
        next_row = rows[next_idx]
        
        # Select and scroll to
        self.table.selectRow(next_row)
        self.table.scrollTo(self.proxy_model.index(next_row, 0))
    
    def keyPressEvent(self, event):
        """[CUSTOM] 2026-04-03 - Keyboard shortcuts for search"""
        if event.key() == Qt.Key.Key_F3:
            # Previous result
            self._search_prev()
        elif event.key() == Qt.Key.Key_F4:
            # Next result
            self._search_next()
        elif event.key() == Qt.Key.Key_Escape:
            # Clear search
            self.search_input.clear()
        else:
            super().keyPressEvent(event)
    
    def _setup_search_completer(self):
        """[CUSTOM] 2026-04-03 - Setup autocomplete for search"""
        from PyQt6.QtWidgets import QCompleter
        self.completer = QCompleter()
        self.completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.completer.setFilterMode(Qt.MatchFlag.MatchContains)
        self.completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        
        # Use model for completions
        self.completion_model = QStandardItemModel()
        self.completer.setModel(self.completion_model)
        
        # [CUSTOM] 2026-04-03 - Ajuster automatiquement selon le contenu
        # Pas de taille fixe - le popup s'adapte automatiquement
        
        self.search_input.setCompleter(self.completer)
    
    def _update_completer(self, text):
        """Update autocomplete suggestions based on current search"""
        if not text or len(text) < 2:
            self.completion_model.clear()
            return
        
        suggestions = set()
        text_lower = text.lower()
        
        # Collect unique values from all columns
        for row in range(self.model.rowCount()):
            for col in range(self.model.columnCount()):
                val = self.model.item(row, col)
                if val:
                    val_text = val.text()
                    if text_lower in val_text.lower():
                        suggestions.add(val_text[:50])  # Limit length
        
        # Update completion model
        self.completion_model.clear()
        for suggestion in sorted(suggestions)[:20]:  # Max 20 suggestions
            self.completion_model.appendRow(QStandardItem(suggestion))

    def _sync_column_widths(self, index, old_size, new_size):
        self.footer.setColumnWidth(index, new_size)

    def _on_data_changed(self):
        if self._auto_fit_enabled:
            self.table.resizeColumnsToContents()
            for i in range(self.model.columnCount()):
                self._sync_column_widths(i, 0, self.table.columnWidth(i))
        
        # معادلة الترتيب المركزية: الأقدم (1) دائماً في الأعلى
        if self.table.isSortingEnabled() and self.model.rowCount() > 0:
            self.table.sortByColumn(0, Qt.SortOrder.AscendingOrder)
            
        self._update_footer()

    def _update_footer(self):
        """Met à jour la barre de pied de tableau avec totaux et compte"""
        self.footer_model.setColumnCount(self.model.columnCount())
        
        # Charger les couleurs du thème
        try:
            theme_id = self.settings_service.get_setting("active_theme", "emerald")
            theme = THEMES.get(theme_id, THEMES["emerald"])
            text_secondary = theme["colors"].get("text_secondary", "#7d8590")
        except:
            text_secondary = "#7d8590"
        
        for col in range(self.model.columnCount()):
            if self.table.isColumnHidden(col):
                self.footer.setColumnHidden(col, True)
                continue
            self.footer.setColumnHidden(col, False)
            
            # [UNIFIED] 2026-04-08 - Auto-detect alignment if not set
            align_type = self._header_align_map.get(col)
            
            # If no alignment set, auto-detect from header name
            if align_type is None:
                header = self.model.horizontalHeaderItem(col)
                header_name = header.text().lower() if header else ""
                
                if col == 0 and self._row_numbering_enabled:
                    align_type = 'center'
                elif any(kw in header_name for kw in ['montant', 'solde', 'prix', 'total', 'qte', 'nombre', 'amount', 'balance', 'credit', 'debit']):
                    align_type = 'right'
                elif any(kw in header_name for kw in ['date', 'taux', 'téléphone', 'phone', 'référence', 'réf']):
                    align_type = 'center'
                else:
                    align_type = 'left'
            
            # N° column - afficher le count - ALIGNEMENT CENTRE
            if col == 0 and self._row_numbering_enabled:
                count = self.proxy_model.rowCount()
                item = QStandardItem(str(count))  # [UNIFIED] 2026-04-08 - Afficher uniquement le nombre
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                item.setForeground(QColor(text_secondary))
                self.footer_model.setItem(0, col, item)
            
            # Colonnes numériques (amount, right) - calculer la somme - ALIGNEMENT DROITE
            elif align_type == 'right':
                total = 0.0
                numeric_count = 0
                for row in range(self.proxy_model.rowCount()):
                    val_str = str(self.proxy_model.index(row, col).data())
                    # Extraire seulement les nombres
                    clean_val = re.sub(r'[^\d.,-]', '', val_str).replace(',', '.')
                    try:
                        if clean_val and clean_val != "." and clean_val != "-":
                            total += float(clean_val)
                            numeric_count += 1
                    except ValueError:
                        continue

                # [UNIFIED] 2026-04-08 - Afficher somme seulement si la colonne contient des nombres réels
                if numeric_count > 0 and total != 0:
                    item = QStandardItem(f"{total:,.2f}")
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    item.setForeground(QColor(text_secondary))
                    self.footer_model.setItem(0, col, item)
                else:
                    self.footer_model.setItem(0, col, QStandardItem(""))
            
            # Colonnes de texte (names, descriptions) - ne pas sommer - ALIGNEMENT GAUCHE
            else:
                self.footer_model.setItem(0, col, QStandardItem(""))
            
            self.footer.setColumnWidth(col, self.table.columnWidth(col))

    def _on_selection_changed(self):
        selected_rows = self.get_selected_rows()
        has_selection = len(selected_rows) > 0
        self.edit_action.setEnabled(has_selection)
        self.delete_action.setEnabled(has_selection)
        self.selectionChanged.emit(selected_rows)

    def _on_row_double_clicked(self, index):
        if self.edit_action.isEnabled():
            self.editClicked.emit(index.row())

    def set_headers(self, headers: list, align_map: dict = None):
        """
        Configure les en-têtes avec alignement personnalisé.
        
        Args:
            headers: Liste des noms des colonnes
            align_map: Dict {index: 'left'|'center'|'right'|'amount'} pour spécifier l'alignement
                      Ex: {0: 'center', 2: 'amount', 3: 'left'}
        """
        self.model.setHorizontalHeaderLabels(headers)
        self.footer_model.setColumnCount(len(headers))
        
        # Store align map for later use in add_row and footer
        self._header_align_map = align_map or {}
        
        # [UNIFIED] 2026-04-08 - Auto-detect alignment based on header name
        # N° or # → center
        # Amount columns (Montant, Solde, Prix, Total, Qte, Nombre) → right
        # Date → center
        # Everything else → left
        for idx, header in enumerate(headers):
            if idx in self._header_align_map:
                continue  # Skip if already specified
            
            header_lower = header.lower()
            
            if header in ["N°", "N", "#", "Numéro"] or "n°" in header_lower or "#" in header:
                self._header_align_map[idx] = 'center'
            elif any(kw in header_lower for kw in ['montant', 'solde', 'prix', 'total', 'qte', 'nombre', 'amount', 'balance', 'credit', 'debit']):
                self._header_align_map[idx] = 'right'
            elif any(kw in header_lower for kw in ['date', 'taux', 'téléphone', 'phone', 'référence', 'réf']):
                self._header_align_map[idx] = 'center'
            else:
                self._header_align_map[idx] = 'left'
        
        # Détecter si le premier col est N° pour activer le numérotage auto
        if headers and headers[0] == "N°":
            self._row_numbering_enabled = True
            # Par défaut, N° est centré
            self._header_align_map[0] = 'center'
        else:
            self._row_numbering_enabled = False
        
        # [UNIFIED] 2026-04-08 - Auto-hide ID column (column 1) for all tables
        if len(headers) > 1:
            self.hide_column(1)
        
        QTimer.singleShot(100, self._on_data_changed)

    def add_row(self, data: list, color: str = None, is_active: bool = True):
        """Ajoute une ligne. Si N° est activé, insère l'index auto formaté."""
        if self._row_numbering_enabled:
            data_copy = list(data)
            # Stocker le numéro formaté pour l'affichage: 1
            # Et utiliser UserRole pour le tri: 000001
            row_num = self.model.rowCount() + 1
            data_copy[0] = str(row_num)  # Affichage: 1
            row_to_add = data_copy
        else:
            row_to_add = data
            
        items = []
        for col_idx, value in enumerate(row_to_add):
            if isinstance(value, QStandardItem):
                item = value
                if not item.isCheckable():
                    item.setEditable(False)
            else:
                val_str = str(value) if value is not None else ""
                item = QStandardItem(val_str)
                item.setEditable(False)
                
                # [CUSTOM] 2026-04-03 - N° column: display 1 but sort properly
                if col_idx == 0 and self._row_numbering_enabled:
                    # Pour le tri: stocker la valeur numérique dans UserRole
                    item.setData(self.model.rowCount() + 1, Qt.ItemDataRole.UserRole)
            
            # --- ALIGNEMENT UNIFIÉ ---
            align_type = self._header_align_map.get(col_idx, 'auto')  # [FIX] 2026-04-04 - défaut 'auto' pour activer la détection automatique

            # Auto-détection améliorée si pas spécifié
            if align_type == 'auto':
                if col_idx == 0 and self._row_numbering_enabled:
                    align_type = 'center'
                elif 'DA' in val_str or 'EUR' in val_str or 'USD' in val_str or '$' in val_str or '€' in val_str:
                    align_type = 'amount'
                elif '%' in val_str:
                    align_type = 'percentage'
                elif self._is_date(val_str):
                    align_type = 'date'
                elif self._is_status(val_str):
                    align_type = 'status'
                elif self._is_code(val_str):
                    align_type = 'code'
                elif self._is_phone(val_str):
                    align_type = 'phone'
                else:
                    # Vérifier si c'est un nombre
                    clean = val_str.replace(' ', '').replace(',', '').replace('.', '')
                    if clean.replace('-', '').isdigit():
                        align_type = 'amount'
                    else:
                        align_type = 'left'

            # Appliquer l'alignement selon le type
            if align_type == 'center':
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            elif align_type == 'right' or align_type == 'amount' or align_type == 'percentage':
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            elif align_type == 'date' or align_type == 'status' or align_type == 'code' or align_type == 'phone':
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            else:  # left
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            
            # [UNIFIED] 2026-04-08 - Appliquer couleurs soldes depuis settings
            # Détecter si la valeur est un montant (contient DA, EUR, USD ou est un nombre)
            if align_type == 'right' or align_type == 'amount':
                # Extraire le nombre pour déterminer le signe
                clean = val_str.replace(' ', '').replace(',', '.').replace('DA', '').replace('EUR', '').replace('USD', '').replace('$', '').replace('€', '')
                try:
                    amount = float(clean) if clean and clean != '.' and clean != '-' else 0
                    # Charger les couleurs depuis settings
                    color_pos = self.settings_service.get_setting("color_positive_balance", "#238636")
                    color_neg = self.settings_service.get_setting("color_negative_balance", "#f85149")
                    
                    if amount > 0:
                        item.setForeground(QColor(color_pos))
                    elif amount < 0:
                        item.setForeground(QColor(color_neg))
                except (ValueError, AttributeError):
                    pass
            
            # Styles (Actif/Inactif/Couleur)
            if not is_active:
                item.setForeground(QColor("#8b949e"))
                font = item.font()
                font.setItalic(True)
                item.setFont(font)
            elif color:
                item.setForeground(QColor(color))
                
            items.append(item)
            
        self.model.appendRow(items)
        return self.model.rowCount() - 1

    def set_row_background_color(self, row: int, color: str):
        """Définit la couleur de fond pour toute une ligne"""
        for col in range(self.model.columnCount()):
            item = self.model.item(row, col)
            if item:
                item.setBackground(QColor(color))

    def set_row_foreground_color(self, row: int, color: str):
        """Définit la couleur de texte pour toute une ligne"""
        for col in range(self.model.columnCount()):
            item = self.model.item(row, col)
            if item:
                item.setForeground(QColor(color))

    # ========================================================================
    # AUTO-DETECTION HELPERS (Tree System Alignment)
    # ========================================================================

    @staticmethod
    def _is_date(val: str) -> bool:
        """Détecte si une valeur ressemble à une date"""
        import re
        # Formats: 2026-04-04, 04/04/2026, 04-04-2026, 2026/04/04
        date_patterns = [
            r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}',  # YYYY-MM-DD ou YYYY/MM/DD
            r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}',  # DD/MM/YYYY ou DD-MM-YYYY
            r'^\d{1,2}\s+(Jan|Fév|Mar|Avr|Mai|Jun|Jul|Aoû|Sep|Oct|Nov|Déc)',  # 4 Avr 2026
        ]
        return any(re.search(p, val, re.IGNORECASE) for p in date_patterns)

    @staticmethod
    def _is_status(val: str) -> bool:
        """Détecte si une valeur ressemble à un statut"""
        status_values = {
            'actif', 'inactif', 'active', 'inactive', 'supprimé', 'supprime',
            'validée', 'validee', 'en_attente', 'en cours', 'encours',
            'annulée', 'annulee', 'ouvert', 'fermé', 'ferme', 'open', 'closed',
            'oui', 'non', 'yes', 'no', 'principal', 'cash', 'credit', 'crédit',
            'chèque', 'cheque', 'virement', 'traite', 'effet', 'especes', 'espèces',
            'cours', 'payé', 'paye', 'impayé', 'impaye', 'partiel', 'partial',
            'livré', 'livre', 'en transit', 'transit', 'recu', 'reçu', 'retour'
        }
        return val.strip().lower() in status_values

    @staticmethod
    def _is_code(val: str) -> bool:
        """Détecte si une valeur ressemble à un code/identifiant"""
        import re
        # Formats: MAIN_DZD, CAISSE, EUR, CNY, ABC123, CODE-123
        code_patterns = [
            r'^[A-Z]{2,}[_\-][A-Z0-9]+$',  # MAIN_DZD, CODE-123
            r'^[A-Z]{3}$',  # EUR, USD, CNY
            r'^[A-Z]+[_\-]?\d+$',  # CAISSE1, COMPTE_1
            r'^[A-Z]{2,}$',  # CAISSE, COMPTE, BANQUE (majuscules uniquement)
        ]
        return any(re.match(p, val.strip()) for p in code_patterns)

    @staticmethod
    def _is_phone(val: str) -> bool:
        """Détecte si une valeur ressemble à un numéro de téléphone"""
        import re
        # Formats: 0555123456, +213 555 123 456, 00-213-555-123-456
        phone_patterns = [
            r'^\+?\d{1,3}[\s\-]?\d{2,4}[\s\-]?\d{2,4}[\s\-]?\d{2,4}$',
            r'^0\d{8,9}$',  # Format algérien: 0XXXXXXXXX
            r'^00\d{10,15}$',  # Format international: 00XXXXXXXXXX
        ]
        clean = val.replace(' ', '').replace('-', '').replace('.', '')
        return any(re.match(p, clean) for p in phone_patterns)

    def get_selected_rows(self) -> list:
        selection = self.table.selectionModel()
        if not selection.hasSelection(): return []
        proxy_indexes = selection.selectedRows()
        return [self.proxy_model.mapToSource(idx).row() for idx in proxy_indexes]

    def get_row_data(self, row: int) -> list:
        return [self.model.item(row, col).text() if self.model.item(row, col) else "" for col in range(self.model.columnCount())]

    def clear_rows(self):
        self.model.setRowCount(0)

    def hide_column(self, column: int):
        if column not in self._permanent_hidden_columns: self._permanent_hidden_columns.append(column)
        self.table.setColumnHidden(column, True)
        self._update_footer()

    def resize_columns_to_contents(self):
        self.table.resizeColumnsToContents()
        for i in range(self.model.columnCount()):
            self._sync_column_widths(i, 0, self.table.columnWidth(i))

    def add_action_button(self, text: str, icon: str, callback):
        btn = QPushButton(f" {text}")
        if icon:
            btn.setIcon(IconManager.get_icon(icon, color="#ffffff", size=16))
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setStyleSheet("""
            QPushButton {
                font-weight: bold;
                font-size: 11px;
                background-color: #1f6feb;
                color: white;
                border: 1px solid #145a9e;
                border-radius: 6px;
                padding: 5px 12px;
            }
            QPushButton:hover { background-color: #388bfd; }
            QPushButton:pressed { background-color: #1158c7; }
        """)
        btn.clicked.connect(callback)
        self.toolbar.addWidget(btn)

    def update_actions_for_status(self, status: str):
        self.add_action.setEnabled(True)
        self.edit_action.setEnabled(True)
        self.delete_action.setVisible(status == "active")
        self.delete_action.setEnabled(status == "active")
        self.restore_action.setVisible(status == "inactive")
        self.restore_action.setEnabled(status == "inactive")

    def set_data(self, data: list):
        """Remplace toutes les données et gère le Running Balance si activé"""
        self.clear_rows()
        self.table.setSortingEnabled(False)
        
        current_bal = self._initial_balance
        for row_data in data:
            row_copy = list(row_data)
            # Calcul du running balance si l'index est défini
            if self._running_balance_col >= 0:
                # On assume que le montant est dans une colonne et qu'on traite Crédit/Débit
                # Mais simplifions : si le view a déjà géré le calcul, on passe.
                # Ici on peut recalculer si besoin. Pour l'instant on fait juste add_row.
                pass
            self.add_row(row_copy)
            
        self.table.setSortingEnabled(True)
        self._on_data_changed()

    def set_running_balance(self, column_index: int, initial: float = 0.0):
        """Active le calcul du solde progressif sur une colonne"""
        self._running_balance_col = column_index
        self._initial_balance = initial

    def export_to_pdf(self):
        """Exporte le tableau en PDF avec le style Premium"""
        from modules.settings.service import SettingsService
        settings_service = SettingsService()
        
        dialog = ExportPreviewDialog(table_id=self.table_id or "Rapport", parent=self)
        dialog.setWindowTitle("📄 Exportation PDF Premium")
        dialog.set_button_text("🚀 Générer le PDF")
        
        # Store reference for PDF preview
        self._current_html = None
        self._current_settings = settings_service
        
        def generate_preview():
            opts = dialog.get_options()
            return self._generate_premium_html(opts, settings_service)
        
        dialog.set_pdf_callback(generate_preview)
        
        if not dialog.exec(): return
        
        opts = dialog.get_options()
        html = self._generate_premium_html(opts, settings_service)
        
        # Generate default filename from title
        default_name = opts['title'].replace(" ", "_").replace(":", "")[:50]
        default_filename = f"{default_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        filename, _ = QFileDialog.getSaveFileName(
            self, 
            "Enregistrer le PDF", 
            default_filename, 
            "PDF (*.pdf)"
        )
        if not filename: return
        
        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            try:
                printer.setPageOrientation(QPageLayout.Orientation.Landscape)
            except: pass
            printer.setOutputFileName(filename)
            
            doc = QTextDocument()
            doc.setHtml(html)
            doc.print(printer)
            QMessageBox.information(self, "Succès", "Document PDF Premium généré avec succès!")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Échec PDF : {str(e)}")

    def _create_toolbar(self) -> QToolBar:
        toolbar = QToolBar()
        toolbar.setObjectName("tableToolbar")
        toolbar.setMovable(False)
        toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
        
        # [CUSTOM] تنسيق أزرار شريط الأدوات
        # [WHY]: ToolButtonTextOnly بدون تنسيق يُظهر أزراراً بدون إطار مميز.
        #        هذا التنسيق يعيد الإطار (Border) ليبقى البرنامج احترافياً ومتناسقاً.
        # [DATE]: 2026-03-29
        toolbar.setStyleSheet("""
            QToolBar {
                border: none;
                spacing: 3px;
                padding: 4px 6px;
            }
            QToolBar QToolButton {
                border: 1px solid rgba(255,255,255,0.12);
                border-radius: 5px;
                padding: 4px 12px;
                margin: 1px 2px;
                font-weight: bold;
                font-size: 12px;
                min-width: 60px;
            }
            QToolBar QToolButton:hover {
                border-color: rgba(255,255,255,0.35);
                background-color: rgba(255,255,255,0.08);
            }
            QToolBar QToolButton:pressed {
                background-color: rgba(255,255,255,0.15);
            }
            QToolBar::separator {
                width: 1px;
                background: rgba(255,255,255,0.1);
                margin: 4px 3px;
            }
        """)
        
        self.add_action = QAction("Nouveau", self)
        self.add_action.triggered.connect(self.addClicked.emit)
        toolbar.addAction(self.add_action)
        
        self.edit_action = QAction("Modifier", self)
        self.edit_action.triggered.connect(lambda: self.editClicked.emit(self.get_selected_rows()[0]) if self.get_selected_rows() else None)
        self.edit_action.setEnabled(False)
        toolbar.addAction(self.edit_action)
        
        self.delete_action = QAction("Supprimer", self)
        self.delete_action.triggered.connect(lambda: self.deleteClicked.emit(self.get_selected_rows()[0]) if self.get_selected_rows() else None)
        self.delete_action.setEnabled(False)
        toolbar.addAction(self.delete_action)
        
        self.restore_action = QAction("Restaurer", self)
        self.restore_action.triggered.connect(lambda: self.restoreClicked.emit(self.get_selected_rows()[0]) if self.get_selected_rows() else None)
        self.restore_action.setVisible(False)
        self.restore_action.setEnabled(False)
        toolbar.addAction(self.restore_action)
        
        toolbar.addSeparator()
        
        self.refresh_action = QAction("Rafraîchir", self)
        self.refresh_action.triggered.connect(self.refreshClicked.emit)
        toolbar.addAction(self.refresh_action)
        
        toolbar.addSeparator()
        
        from utils.icon_manager import IconManager

        self.pdf_action = QAction(IconManager.get_icon("file-text", size=20), "", self)
        self.pdf_action.setToolTip("Exporter en PDF")
        self.pdf_action.triggered.connect(self.export_to_pdf)
        toolbar.addAction(self.pdf_action)

        self.export_action = QAction(IconManager.get_icon("file-spreadsheet", size=20), "", self)
        self.export_action.setToolTip("Exporter vers Excel")
        self.export_action.triggered.connect(self._export_to_excel)
        toolbar.addAction(self.export_action)

        self.print_action = QAction(IconManager.get_icon("printer", size=20), "", self)
        self.print_action.setToolTip("Imprimer")
        self.print_action.triggered.connect(self._on_print_preview_clicked)
        toolbar.addAction(self.print_action)

        for action in [self.pdf_action, self.export_action, self.print_action]:
            btn = toolbar.widgetForAction(action)
            if btn:
                btn.setProperty("icon_only", True)
                btn.setStyleSheet("min-width: 32px; padding: 4px 6px;")
                btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonIconOnly)
        
        toolbar.addSeparator()

        return toolbar

    def _show_context_menu(self, position):
        menu = QMenu()

        # Actions principales
        menu.addAction(self.add_action)
        if self.edit_action.isEnabled(): menu.addAction(self.edit_action)
        if self.delete_action.isVisible() and self.delete_action.isEnabled(): menu.addAction(self.delete_action)
        if self.restore_action.isVisible() and self.restore_action.isEnabled(): menu.addAction(self.restore_action)

        menu.addSeparator()

        # Copier
        copy_menu = menu.addMenu("Copier")
        copy_menu.addAction("Cellule", lambda: QApplication.clipboard().setText(str(self.table.currentIndex().data())))
        
        # [UNIFIED] 2026-04-08 - Copy row and copy table
        copy_menu.addAction("Ligne", lambda: self._copy_row_to_clipboard())
        copy_menu.addAction("Tableau", lambda: self._copy_table_to_clipboard())

        menu.addSeparator()

        # Export & Import
        menu.addAction(self.pdf_action)
        menu.addAction(self.export_action)
        menu.addAction(self.print_action)

        menu.addSeparator()

        # Rafraîchir
        menu.addAction(self.refresh_action)

        menu.addSeparator()

        # Sous-menu: Colonnes
        columns_menu = menu.addMenu("Colonnes")
        col_count = self.model.columnCount()
        for i in range(col_count):
            if i in self._permanent_hidden_columns:
                continue
            header = self.model.headerData(i, Qt.Orientation.Horizontal) or f"Col {i}"
            action = columns_menu.addAction(str(header))
            action.setCheckable(True)
            action.setChecked(not self.table.isColumnHidden(i))
            action.triggered.connect(lambda checked, col=i: self.table.setColumnHidden(col, not checked))

        menu.exec(self.table.viewport().mapToGlobal(position))

    def _copy_row_to_clipboard(self):
        """[UNIFIED] 2026-04-08 - Copy selected row to clipboard (tab-separated)"""
        current_row = self.table.currentIndex().row()
        if current_row < 0:
            return
        
        row_data = []
        for col in range(self.model.columnCount()):
            if self.table.isColumnHidden(col):
                continue
            index = self.model.index(current_row, col)
            value = index.data() or ""
            row_data.append(str(value))
        
        # Copy as tab-separated (for Excel/Google Sheets paste)
        clipboard_text = "\t".join(row_data)
        QApplication.clipboard().setText(clipboard_text)

    def _copy_table_to_clipboard(self):
        """[UNIFIED] 2026-04-08 - Copy entire table to clipboard (Excel format)"""
        lines = []
        
        # Header row
        headers = []
        for col in range(self.model.columnCount()):
            if self.table.isColumnHidden(col):
                continue
            header = self.model.headerData(col, Qt.Orientation.Horizontal) or ""
            headers.append(str(header))
        lines.append("\t".join(headers))
        
        # Data rows (only visible rows via proxy model)
        for row in range(self.proxy_model.rowCount()):
            row_data = []
            for col in range(self.model.columnCount()):
                if self.table.isColumnHidden(col):
                    continue
                index = self.proxy_model.index(row, col)
                value = index.data() or ""
                row_data.append(str(value))
            lines.append("\t".join(row_data))
        
        # Copy all to clipboard
        clipboard_text = "\n".join(lines)
        QApplication.clipboard().setText(clipboard_text)

    def _export_to_excel(self):
        if not OPENPYXL_AVAILABLE:
            self._export_to_csv_legacy()
            return

        # 1. Préparer le dialogue de configuration
        from modules.settings.service import SettingsService
        settings_service = SettingsService()
        
        dialog = ExportPreviewDialog(table_id=self.table_id or "Rapport", parent=self)
        if not dialog.exec(): return # Annulé
        
        opts = dialog.get_options()
        
        # 2. Demander l'emplacement de sauvegarde
        default_name = opts['title'].replace(" ", "_").replace(":", "")[:50]
        default_filename = f"{default_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filename, _ = QFileDialog.getSaveFileName(
            self, 
            "Exporter vers Excel (.xlsx)", 
            default_filename, 
            "Excel Files (*.xlsx)"
        )
        if not filename: return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapport"
            
            # --- ZONE PREMIUM : EN-TETE DE L'ENTREPRISE ---
            current_row = 1
            if opts["include_company"]:
                comp_name = settings_service.get_setting("company_name", "Ma Société")
                addr = settings_service.get_setting("company_address", "")
                tel = settings_service.get_setting("company_phone", "")
                
                # Nom de l'entreprise (Gros & Gras)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                cell_name = ws.cell(row=current_row, column=1, value=comp_name.upper())
                cell_name.font = Font(bold=True, size=16, color="1f6feb")
                current_row += 1
                
                # Adresse & Tel
                if addr or tel:
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                    ws.cell(row=current_row, column=1, value=f"📍 {addr} | 📞 {tel}").font = Font(italic=True, color="444444")
                    current_row += 1
                
                current_row += 1 # Espace
            
            # --- ZONE PREMIUM : TITRE DU RAPPORT ---
            # On fusionne sur toute la largeur du tableau
            num_cols = len([c for c in range(self.model.columnCount()) if not self.table.isColumnHidden(c)])
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
            title_cell = ws.cell(row=current_row, column=1, value=opts["title"])
            title_cell.font = Font(bold=True, size=20)
            title_cell.alignment = Alignment(horizontal="center")
            current_row += 1
            
            if opts["subtitle"]:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
                sub_cell = ws.cell(row=current_row, column=1, value=opts["subtitle"])
                sub_cell.font = Font(italic=True, size=12, color="666666")
                sub_cell.alignment = Alignment(horizontal="center")
                current_row += 1
            
            if opts["include_date"]:
                from datetime import datetime
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
                dt_cell = ws.cell(row=current_row, column=1, value=f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
                dt_cell.font = Font(size=9, color="999999")
                dt_cell.alignment = Alignment(horizontal="right")
                current_row += 1
            
            current_row += 1 # Espace avant le tableau
            table_start_row = current_row
            
            # 3. EN-TETES DU TABLEAU
            headers = [self.proxy_model.headerData(c, Qt.Orientation.Horizontal) for c in range(self.model.columnCount()) if not self.table.isColumnHidden(c)]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=table_start_row, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1f6feb", end_color="1f6feb", fill_type="solid") # Bleu Premium
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Borders
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # 4. DONNEES DU TABLEAU
            data_row_idx = table_start_row + 1
            for r in range(self.proxy_model.rowCount()):
                col_idx_excel = 1
                for c in range(self.model.columnCount()):
                    if not self.table.isColumnHidden(c):
                        val = self.proxy_model.index(r, c).data()
                        # Nettoyage des nombres
                        processed_val = val
                        if val and isinstance(val, str):
                            clean_val = val.replace(" ", "").replace("DA", "").replace("$", "").replace("€", "").replace(",", ".")
                            try:
                                processed_val = float(clean_val)
                            except ValueError: pass
                        
                        cell = ws.cell(row=data_row_idx, column=col_idx_excel, value=processed_val)
                        
                        # Style de la cellule
                        cell.alignment = Alignment(vertical="center")
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        # Style Zèbre
                        if opts["zebra_stripes"] and (data_row_idx % 2 == 0):
                            cell.fill = PatternFill(start_color="f1f8ff", end_color="f1f8ff", fill_type="solid")
                        
                        col_idx_excel += 1
                data_row_idx += 1

            # 5. AUTO-FILTER & FREEZE
            if opts["auto_filters"]:
                ws.auto_filter.ref = f"{get_column_letter(1)}{table_start_row}:{get_column_letter(num_cols)}{data_row_idx-1}"
            ws.freeze_panes = f"A{table_start_row + 1}"
            
            # 6. AUTO-AJUSTEMENT DES COLONNES
            col_idx = 1
            for col_cells in ws.iter_cols(min_row=table_start_row, max_row=data_row_idx-1, min_col=1, max_col=num_cols):
                max_length = 0
                column = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = min(max_length + 4, 50) # Max 50 chars width
                col_idx += 1

            wb.save(filename)
            QMessageBox.information(self, "Succès", "Rapport Premium généré avec succès!")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Échec de l'exportation : {str(e)}")

    def _export_to_csv_legacy(self):
        default_filename = f"export_{datetime.now().strftime('%Y%m%d')}.csv"
        filename, _ = QFileDialog.getSaveFileName(
            self, 
            "Exporter vers Excel (CSV)", 
            default_filename, 
            "CSV compatible Excel (*.csv)"
        )
        if not filename: return
        try:
            with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow([self.proxy_model.headerData(c, Qt.Orientation.Horizontal) for c in range(self.model.columnCount()) if not self.table.isColumnHidden(c)])
                for r in range(self.proxy_model.rowCount()):
                    writer.writerow([str(self.proxy_model.index(r, c).data()) for c in range(self.model.columnCount()) if not self.table.isColumnHidden(c)])
            QMessageBox.information(self, "Succès", "Exportation CSV réussie!")
        except Exception as e: QMessageBox.critical(self, "Erreur", f"Échec CSV : {str(e)}")

    def _on_print_preview_clicked(self):
        """Lance l'aperçu avant impression avec configuration préalable"""
        from modules.settings.service import SettingsService
        settings_service = SettingsService()
        
        dialog = ExportPreviewDialog(table_id=self.table_id or "Rapport", parent=self)
        dialog.setWindowTitle("🖨️ Configuration de l'impression")
        dialog.set_button_text("👁️ Afficher l'aperçu (Plein écran)")
        
        # Store reference for PDF preview
        self._current_html = None
        self._current_settings = settings_service
        
        def generate_preview():
            opts = dialog.get_options()
            return self._generate_premium_html(opts, settings_service)
        
        dialog.set_pdf_callback(generate_preview)
        
        if not dialog.exec(): return
        
        opts = dialog.get_options()
        html = self._generate_premium_html(opts, settings_service)
        
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        try:
            printer.setPageOrientation(QPageLayout.Orientation.Landscape)
        except: pass
        
        preview = QPrintPreviewDialog(printer, self)
        preview.setWindowTitle(f"Aperçu : {opts['title']}")
        preview.setMinimumSize(1000, 800)
        
        def handle_print(p):
            doc = QTextDocument()
            doc.setHtml(html)
            doc.print(p)
            
        preview.paintRequested.connect(handle_print)
        preview.exec()

    def _generate_premium_html(self, opts, settings_service):
        """Génère le code HTML stylisé pour l'impression / PDF"""
        comp_name = settings_service.get_setting("company_name", "Ma Société")
        addr = settings_service.get_setting("company_address", "")
        tel = settings_service.get_setting("company_phone", "")
        
        from datetime import datetime
        now = datetime.now().strftime("%d/%m/%Y à %H:%M")
        
        # Style CSS pour l'impression
        css = """
        <style>
            body { font-family: 'Segoe UI', Arial, sans-serif; color: #333; margin: 20px; }
            .header { border-bottom: 2px solid #1f6feb; padding-bottom: 15px; margin-bottom: 30px; }
            .company-name { font-size: 22pt; font-weight: bold; color: #1f6feb; text-transform: uppercase; }
            .company-info { font-size: 10pt; color: #555; }
            .report-title-box { text-align: center; margin-bottom: 25px; }
            .report-title { font-size: 28pt; font-weight: bold; margin: 0; }
            .report-subtitle { font-size: 14pt; color: #666; font-style: italic; }
            .metadata { text-align: right; font-size: 9pt; color: #888; margin-bottom: 10px; }
            
            table { width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 10pt; margin-left: auto; margin-right: auto; }
            th { background-color: #1f6feb; color: white; padding: 10px; font-weight: bold; border: 1px solid #145a9e; text-align: center; }
            td { padding: 8px; border: 1px solid #ddd; vertical-align: middle; }
            .zebra { background-color: #f8faff; }
            .text-right { text-align: right; }
            .text-center { text-align: center; }
            .footer { margin-top: 30px; font-size: 8pt; text-align: center; border-top: 1px solid #eee; padding-top: 10px; }
        </style>
        """
        
        html = f"<html><head>{css}</head><body>"
        
        # 1. En-tête Entreprise
        if opts["include_company"]:
            html += f'<div class="header">'
            html += f'<div class="company-name">{comp_name}</div>'
            html += f'<div class="company-info">{addr} | Tél: {tel}</div>'
            html += '</div>'
            
        # 2. Titre du rapport
        html += f'<div class="report-title-box">'
        html += f'<h1 class="report-title">{opts["title"]}</h1>'
        if opts["subtitle"]:
            html += f'<div class="report-subtitle">{opts["subtitle"]}</div>'
        html += '</div>'
        
        # 3. Métadonnées
        if opts["include_date"]:
            html += f'<div class="metadata">Généré le {now}</div>'
            
        # 4. Le Tableau
        headers = [self.proxy_model.headerData(c, Qt.Orientation.Horizontal) for c in range(self.model.columnCount()) if not self.table.isColumnHidden(c)]
        
        html += '<table><thead><tr>'
        for h in headers:
            html += f'<th>{h}</th>'
        html += '</tr></thead><tbody>'
        
        for r in range(self.proxy_model.rowCount()):
            classes = "zebra" if (opts["zebra_stripes"] and r % 2 != 0) else ""
            html += f'<tr class="{classes}">'
            for c in range(self.model.columnCount()):
                if not self.table.isColumnHidden(c):
                    val = self.proxy_model.index(r, c).data()
                    val_str = str(val) if val is not None else ""
                    
                    # Alignements
                    align_class = ""
                    if any(s in val_str for s in ["DA", "$", "€", "USD", "EUR"]): align_class = "text-right"
                    elif c == 0: align_class = "text-center" # N°
                    
                    html += f'<td class="{align_class}">{val_str}</td>'
            html += '</tr>'
            
        html += '</tbody></table>'
        
        # 5. Footer
        html += f'<div class="footer">Logiciel de Gestion Logistique | Rapport interactif</div>'
        
        html += "</body></html>"
        return html
